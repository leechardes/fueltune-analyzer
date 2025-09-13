"""
FuelTune Export/Import System

Sistema avançado de exportação e importação de dados em múltiplos formatos
com templates personalizáveis e operações em lote.

Classes:
    ExportImportManager: Gerenciador principal
    DataExporter: Exportadores especializados
    DataImporter: Importadores especializados
    ExportTemplate: Templates de exportação
    BatchProcessor: Processamento em lote

Author: FuelTune Development Team
Version: 1.0.0
"""

import json
import tempfile
import zipfile
from abc import ABC, abstractmethod
from dataclasses import dataclass, field
from datetime import datetime
from enum import Enum
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

import pandas as pd
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill

from ..utils.logger import get_logger
from .events import DataEvent, event_bus
from .notifications import notify_error, notify_info, notify_success

logger = get_logger(__name__)


class ExportFormat(Enum):
    """Formatos de exportação suportados."""

    CSV = "csv"
    EXCEL = "xlsx"
    JSON = "json"
    PDF = "pdf"
    HTML = "html"
    XML = "xml"
    PARQUET = "parquet"
    SQLITE = "sqlite"
    ZIP = "zip"
    PNG = "png"
    SVG = "svg"


class ImportFormat(Enum):
    """Formatos de importação suportados."""

    CSV = "csv"
    EXCEL = "xlsx"
    JSON = "json"
    XML = "xml"
    PARQUET = "parquet"
    SQLITE = "sqlite"


class ExportType(Enum):
    """Tipos de exportação."""

    SESSION_DATA = "session_data"
    ANALYSIS_RESULTS = "analysis_results"
    CHARTS = "charts"
    REPORTS = "reports"
    VEHICLE_DATA = "vehicle_data"
    SYSTEM_CONFIG = "system_config"


@dataclass
class ExportSettings:
    """Configurações de exportação."""

    format: ExportFormat
    include_metadata: bool = True
    include_charts: bool = True
    include_statistics: bool = True
    compress_data: bool = False
    custom_template: Optional[str] = None
    output_path: Optional[Path] = None
    filename_pattern: str = "{type}_{timestamp}_{session_id}"

    # Configurações específicas por formato
    csv_separator: str = ","
    csv_decimal: str = "."
    excel_sheet_names: Dict[str, str] = field(default_factory=dict)
    pdf_orientation: str = "portrait"  # "portrait" ou "landscape"
    pdf_page_size: str = "A4"
    html_template: Optional[str] = None
    json_indent: int = 2

    def get_filename(self, export_type: ExportType, session_id: str = "default") -> str:
        """Gerar nome do arquivo baseado no padrão."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        filename = self.filename_pattern.format(
            type=export_type.value, timestamp=timestamp, session_id=session_id
        )

        return f"{filename}.{self.format.value}"


@dataclass
class ExportResult:
    """Resultado da operação de exportação."""

    success: bool
    file_path: Optional[Path] = None
    file_size: int = 0
    format: Optional[ExportFormat] = None
    export_type: Optional[ExportType] = None
    duration: float = 0.0
    records_exported: int = 0
    error_message: str = ""
    warnings: List[str] = field(default_factory=list)
    metadata: Dict[str, Any] = field(default_factory=dict)


@dataclass
class ImportResult:
    """Resultado da operação de importação."""

    success: bool
    records_imported: int = 0
    format: Optional[ImportFormat] = None
    duration: float = 0.0
    data: Any = None
    error_message: str = ""
    warnings: List[str] = field(default_factory=list)
    validation_errors: List[str] = field(default_factory=list)
    metadata: Dict[str, Any] = field(default_factory=dict)


class DataExporter(ABC):
    """Classe base para exportadores de dados."""

    def __init__(self, format: ExportFormat):
        self.format = format

    @abstractmethod
    def export(self, data: Any, settings: ExportSettings, output_path: Path) -> ExportResult:
        """Exportar dados no formato específico."""

    def validate_data(self, data: Any) -> List[str]:
        """Validar dados antes da exportação."""
        warnings = []

        if data is None:
            warnings.append("Dados são None")
        elif isinstance(data, pd.DataFrame) and data.empty:
            warnings.append("DataFrame está vazio")
        elif isinstance(data, list) and len(data) == 0:
            warnings.append("Lista está vazia")

        return warnings


class CSVExporter(DataExporter):
    """Exportador para formato CSV."""

    def __init__(self):
        super().__init__(ExportFormat.CSV)

    def export(self, data: Any, settings: ExportSettings, output_path: Path) -> ExportResult:
        """Exportar dados como CSV."""
        start_time = datetime.now()

        try:
            warnings = self.validate_data(data)

            if isinstance(data, pd.DataFrame):
                data.to_csv(
                    output_path,
                    index=False,
                    sep=settings.csv_separator,
                    decimal=settings.csv_decimal,
                    encoding="utf-8",
                )
                records_exported = len(data)

            elif isinstance(data, list):
                # Converter lista para DataFrame
                df = pd.DataFrame(data)
                df.to_csv(
                    output_path,
                    index=False,
                    sep=settings.csv_separator,
                    decimal=settings.csv_decimal,
                    encoding="utf-8",
                )
                records_exported = len(data)

            elif isinstance(data, dict):
                # Converter dict para DataFrame
                df = pd.DataFrame([data])
                df.to_csv(
                    output_path,
                    index=False,
                    sep=settings.csv_separator,
                    decimal=settings.csv_decimal,
                    encoding="utf-8",
                )
                records_exported = 1

            else:
                # Dados simples
                with open(output_path, "w", encoding="utf-8") as f:
                    f.write(str(data))
                records_exported = 1

            file_size = output_path.stat().st_size
            duration = (datetime.now() - start_time).total_seconds()

            return ExportResult(
                success=True,
                file_path=output_path,
                file_size=file_size,
                format=self.format,
                duration=duration,
                records_exported=records_exported,
                warnings=warnings,
            )

        except Exception as e:
            duration = (datetime.now() - start_time).total_seconds()

            return ExportResult(
                success=False, format=self.format, duration=duration, error_message=str(e)
            )


class ExcelExporter(DataExporter):
    """Exportador para formato Excel."""

    def __init__(self):
        super().__init__(ExportFormat.EXCEL)

    def export(self, data: Any, settings: ExportSettings, output_path: Path) -> ExportResult:
        """Exportar dados como Excel."""
        start_time = datetime.now()

        try:
            warnings = self.validate_data(data)

            with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
                records_exported = 0

                if isinstance(data, dict):
                    # Múltiplas planilhas
                    for sheet_name, sheet_data in data.items():
                        if isinstance(sheet_data, pd.DataFrame):
                            sheet_data.to_excel(writer, sheet_name=sheet_name, index=False)
                            records_exported += len(sheet_data)
                        else:
                            # Converter para DataFrame
                            df = pd.DataFrame(
                                [sheet_data] if not isinstance(sheet_data, list) else sheet_data
                            )
                            df.to_excel(writer, sheet_name=sheet_name, index=False)
                            records_exported += len(df)

                elif isinstance(data, pd.DataFrame):
                    sheet_name = settings.excel_sheet_names.get("data", "Data")
                    data.to_excel(writer, sheet_name=sheet_name, index=False)
                    records_exported = len(data)

                    # Adicionar formatação
                    self._format_excel_sheet(writer, sheet_name, data)

                    # Adicionar gráficos se solicitado
                    if settings.include_charts:
                        self._add_excel_charts(writer, sheet_name, data)

                else:
                    # Dados simples
                    df = pd.DataFrame([data] if not isinstance(data, list) else data)
                    df.to_excel(writer, sheet_name="Data", index=False)
                    records_exported = len(df)

                # Adicionar metadados se solicitado
                if settings.include_metadata:
                    self._add_metadata_sheet(writer, data, settings)

            file_size = output_path.stat().st_size
            duration = (datetime.now() - start_time).total_seconds()

            return ExportResult(
                success=True,
                file_path=output_path,
                file_size=file_size,
                format=self.format,
                duration=duration,
                records_exported=records_exported,
                warnings=warnings,
            )

        except Exception as e:
            duration = (datetime.now() - start_time).total_seconds()

            return ExportResult(
                success=False, format=self.format, duration=duration, error_message=str(e)
            )

    def _format_excel_sheet(
        self, writer: pd.ExcelWriter, sheet_name: str, data: pd.DataFrame
    ) -> None:
        """Aplicar formatação ao Excel."""
        try:
            worksheet = writer.sheets[sheet_name]

            # Formatação do cabeçalho
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

            for col_num, column in enumerate(data.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            # Ajustar largura das colunas
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter

                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

        except Exception as e:
            logger.warning(f"Erro na formatação Excel: {e}")

    def _add_excel_charts(
        self, writer: pd.ExcelWriter, sheet_name: str, data: pd.DataFrame
    ) -> None:
        """Adicionar gráficos ao Excel."""
        try:
            worksheet = writer.sheets[sheet_name]

            # Procurar colunas numéricas para gráfico
            numeric_columns = data.select_dtypes(include=["number"]).columns.tolist()

            if len(numeric_columns) >= 2:
                # Criar gráfico de linha
                chart = LineChart()
                chart.title = "Dados Numéricos"
                chart.style = 13
                chart.x_axis.title = numeric_columns[0]
                chart.y_axis.title = numeric_columns[1]

                # Definir dados do gráfico
                data_range = Reference(
                    worksheet, min_col=2, min_row=1, max_row=min(len(data) + 1, 100), max_col=3
                )
                chart.add_data(data_range, titles_from_data=True)

                # Adicionar ao worksheet
                worksheet.add_chart(chart, "F5")

        except Exception as e:
            logger.warning(f"Erro ao adicionar gráficos Excel: {e}")

    def _add_metadata_sheet(
        self, writer: pd.ExcelWriter, data: Any, settings: ExportSettings
    ) -> None:
        """Adicionar planilha de metadados."""
        try:
            metadata = [
                ["Exportado em", datetime.now().strftime("%d/%m/%Y %H:%M:%S")],
                ["Formato", settings.format.value],
                ["Registros", len(data) if hasattr(data, "__len__") else "N/A"],
                ["Sistema", "FuelTune Analyzer v1.0"],
            ]

            metadata_df = pd.DataFrame(metadata, columns=["Campo", "Valor"])
            metadata_df.to_excel(writer, sheet_name="Metadata", index=False)

        except Exception as e:
            logger.warning(f"Erro ao adicionar metadados: {e}")


class JSONExporter(DataExporter):
    """Exportador para formato JSON."""

    def __init__(self):
        super().__init__(ExportFormat.JSON)

    def export(self, data: Any, settings: ExportSettings, output_path: Path) -> ExportResult:
        """Exportar dados como JSON."""
        start_time = datetime.now()

        try:
            warnings = self.validate_data(data)

            # Preparar dados para JSON
            if isinstance(data, pd.DataFrame):
                json_data = {
                    "data": data.to_dict(orient="records"),
                    "columns": data.columns.tolist(),
                    "shape": list(data.shape),
                }
                records_exported = len(data)

            elif isinstance(data, (list, dict)):
                json_data = data
                records_exported = len(data) if hasattr(data, "__len__") else 1

            else:
                json_data = {"value": data}
                records_exported = 1

            # Adicionar metadados se solicitado
            if settings.include_metadata:
                json_data = {
                    "metadata": {
                        "exported_at": datetime.now().isoformat(),
                        "format": settings.format.value,
                        "records_count": records_exported,
                        "exporter": "FuelTune Analyzer",
                    },
                    "data": json_data,
                }

            # Escrever arquivo JSON
            with open(output_path, "w", encoding="utf-8") as f:
                json.dump(
                    json_data, f, indent=settings.json_indent, ensure_ascii=False, default=str
                )

            file_size = output_path.stat().st_size
            duration = (datetime.now() - start_time).total_seconds()

            return ExportResult(
                success=True,
                file_path=output_path,
                file_size=file_size,
                format=self.format,
                duration=duration,
                records_exported=records_exported,
                warnings=warnings,
            )

        except Exception as e:
            duration = (datetime.now() - start_time).total_seconds()

            return ExportResult(
                success=False, format=self.format, duration=duration, error_message=str(e)
            )


class DataImporter(ABC):
    """Classe base para importadores de dados."""

    def __init__(self, format: ImportFormat):
        self.format = format

    @abstractmethod
    def import_data(self, file_path: Path, **kwargs) -> ImportResult:
        """Importar dados do arquivo."""

    def validate_file(self, file_path: Path) -> List[str]:
        """Validar arquivo antes da importação."""
        errors = []

        if not file_path.exists():
            errors.append(f"Arquivo não encontrado: {file_path}")
        elif file_path.stat().st_size == 0:
            errors.append("Arquivo está vazio")

        return errors


class CSVImporter(DataImporter):
    """Importador para formato CSV."""

    def __init__(self):
        super().__init__(ImportFormat.CSV)

    def import_data(self, file_path: Path, **kwargs) -> ImportResult:
        """Importar dados de arquivo CSV."""
        start_time = datetime.now()

        try:
            errors = self.validate_file(file_path)
            if errors:
                return ImportResult(success=False, format=self.format, validation_errors=errors)

            # Parâmetros de importação
            separator = kwargs.get("separator", ",")
            decimal = kwargs.get("decimal", ".")
            encoding = kwargs.get("encoding", "utf-8")

            # Importar CSV
            data = pd.read_csv(file_path, sep=separator, decimal=decimal, encoding=encoding)

            duration = (datetime.now() - start_time).total_seconds()

            return ImportResult(
                success=True,
                records_imported=len(data),
                format=self.format,
                duration=duration,
                data=data,
                metadata={
                    "columns": data.columns.tolist(),
                    "shape": list(data.shape),
                    "dtypes": data.dtypes.to_dict(),
                },
            )

        except Exception as e:
            duration = (datetime.now() - start_time).total_seconds()

            return ImportResult(
                success=False, format=self.format, duration=duration, error_message=str(e)
            )


class ExcelImporter(DataImporter):
    """Importador para formato Excel."""

    def __init__(self):
        super().__init__(ImportFormat.EXCEL)

    def import_data(self, file_path: Path, **kwargs) -> ImportResult:
        """Importar dados de arquivo Excel."""
        start_time = datetime.now()

        try:
            errors = self.validate_file(file_path)
            if errors:
                return ImportResult(success=False, format=self.format, validation_errors=errors)

            # Parâmetros de importação
            sheet_name = kwargs.get("sheet_name", 0)  # Primeira planilha por padrão

            # Importar Excel
            if sheet_name == "all":
                # Importar todas as planilhas
                data = pd.read_excel(file_path, sheet_name=None)
                records_imported = sum(len(df) for df in data.values())
            else:
                # Importar planilha específica
                data = pd.read_excel(file_path, sheet_name=sheet_name)
                records_imported = len(data)

            duration = (datetime.now() - start_time).total_seconds()

            metadata = {"file_path": str(file_path)}
            if isinstance(data, pd.DataFrame):
                metadata.update(
                    {
                        "columns": data.columns.tolist(),
                        "shape": list(data.shape),
                        "dtypes": data.dtypes.to_dict(),
                    }
                )

            return ImportResult(
                success=True,
                records_imported=records_imported,
                format=self.format,
                duration=duration,
                data=data,
                metadata=metadata,
            )

        except Exception as e:
            duration = (datetime.now() - start_time).total_seconds()

            return ImportResult(
                success=False, format=self.format, duration=duration, error_message=str(e)
            )


class ExportImportManager:
    """Gerenciador principal de exportação e importação."""

    def __init__(self):
        self.exporters: Dict[ExportFormat, DataExporter] = {
            ExportFormat.CSV: CSVExporter(),
            ExportFormat.EXCEL: ExcelExporter(),
            ExportFormat.JSON: JSONExporter(),
        }

        self.importers: Dict[ImportFormat, DataImporter] = {
            ImportFormat.CSV: CSVImporter(),
            ImportFormat.EXCEL: ExcelImporter(),
        }

        # Diretório padrão para exportações
        self.default_export_dir = Path(tempfile.gettempdir()) / "fueltune_exports"
        self.default_export_dir.mkdir(exist_ok=True)

        logger.info("ExportImportManager inicializado")

    def export_data(
        self,
        data: Any,
        export_type: ExportType,
        settings: ExportSettings,
        session_id: str = "default",
    ) -> ExportResult:
        """Exportar dados."""

        try:
            # Verificar se exporter está disponível
            exporter = self.exporters.get(settings.format)
            if not exporter:
                return ExportResult(
                    success=False,
                    format=settings.format,
                    error_message=f"Exportador não disponível para formato: {settings.format.value}",
                )

            # Determinar caminho de saída
            if settings.output_path:
                output_path = settings.output_path
            else:
                filename = settings.get_filename(export_type, session_id)
                output_path = self.default_export_dir / filename

            # Garantir que diretório existe
            output_path.parent.mkdir(parents=True, exist_ok=True)

            # Notificar início
            notify_info(f"Iniciando exportação: {export_type.value}")

            # Executar exportação
            result = exporter.export(data, settings, output_path)
            result.export_type = export_type

            # Notificar resultado
            if result.success:
                notify_success(
                    f"Exportação concluída: {result.records_exported} registros em {result.duration:.2f}s",
                    file_path=str(result.file_path),
                    file_size=result.file_size,
                )
                self._emit_export_event(result, session_id)
            else:
                notify_error(f"Erro na exportação: {result.error_message}")

            return result

        except Exception as e:
            logger.error(f"Erro no export_data: {e}")
            return ExportResult(
                success=False, format=settings.format, export_type=export_type, error_message=str(e)
            )

    def import_data(
        self, file_path: Union[str, Path], format: ImportFormat, **kwargs
    ) -> ImportResult:
        """Importar dados."""

        try:
            file_path = Path(file_path)

            # Verificar se importer está disponível
            importer = self.importers.get(format)
            if not importer:
                return ImportResult(
                    success=False,
                    format=format,
                    error_message=f"Importador não disponível para formato: {format.value}",
                )

            # Notificar início
            notify_info(f"Iniciando importação: {file_path.name}")

            # Executar importação
            result = importer.import_data(file_path, **kwargs)

            # Notificar resultado
            if result.success:
                notify_success(
                    f"Importação concluída: {result.records_imported} registros em {result.duration:.2f}s"
                )
                self._emit_import_event(result, str(file_path))
            else:
                notify_error(f"Erro na importação: {result.error_message}")

            return result

        except Exception as e:
            logger.error(f"Erro no import_data: {e}")
            return ImportResult(success=False, format=format, error_message=str(e))

    def export_session_data(
        self, session_id: str, data: pd.DataFrame, format: ExportFormat = ExportFormat.CSV
    ) -> ExportResult:
        """Exportar dados de sessão."""

        settings = ExportSettings(format=format, include_metadata=True, include_statistics=True)

        return self.export_data(data, ExportType.SESSION_DATA, settings, session_id)

    def export_analysis_results(
        self, session_id: str, results: Dict[str, Any], format: ExportFormat = ExportFormat.JSON
    ) -> ExportResult:
        """Exportar resultados de análise."""

        settings = ExportSettings(format=format, include_metadata=True, include_charts=True)

        return self.export_data(results, ExportType.ANALYSIS_RESULTS, settings, session_id)

    def export_multiple_sessions(
        self, sessions_data: Dict[str, pd.DataFrame], format: ExportFormat = ExportFormat.EXCEL
    ) -> ExportResult:
        """Exportar múltiplas sessões em um arquivo."""

        settings = ExportSettings(
            format=format,
            include_metadata=True,
            excel_sheet_names={
                session_id: f"Session_{session_id}" for session_id in sessions_data.keys()
            },
        )

        return self.export_data(sessions_data, ExportType.SESSION_DATA, settings, "multiple")

    def create_export_package(
        self,
        session_id: str,
        include_data: bool = True,
        include_analysis: bool = True,
        include_charts: bool = True,
    ) -> ExportResult:
        """Criar pacote completo de exportação."""

        try:
            # Criar diretório temporário para o pacote
            with tempfile.TemporaryDirectory() as temp_dir:
                Path(temp_dir)
                package_files = []

                # TODO: Integrar com dados reais quando disponível
                # Por enquanto, criar arquivos de exemplo

                if include_data:
                    # Exportar dados da sessão
                    dummy_data = pd.DataFrame(
                        {
                            "timestamp": range(100),
                            "rpm": range(800, 900),
                            "throttle": [i * 0.1 for i in range(100)],
                        }
                    )

                    data_result = self.export_session_data(session_id, dummy_data, ExportFormat.CSV)
                    if data_result.success:
                        package_files.append(data_result.file_path)

                if include_analysis:
                    # Exportar resultados de análise
                    analysis_data = {
                        "session_id": session_id,
                        "statistics": {"avg_rpm": 850, "max_rpm": 899},
                        "analysis_timestamp": datetime.now().isoformat(),
                    }

                    analysis_result = self.export_analysis_results(
                        session_id, analysis_data, ExportFormat.JSON
                    )
                    if analysis_result.success:
                        package_files.append(analysis_result.file_path)

                # Criar arquivo ZIP
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                zip_filename = f"fueltune_package_{session_id}_{timestamp}.zip"
                zip_path = self.default_export_dir / zip_filename

                with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zipf:
                    for file_path in package_files:
                        if file_path and file_path.exists():
                            zipf.write(file_path, file_path.name)

                file_size = zip_path.stat().st_size

                return ExportResult(
                    success=True,
                    file_path=zip_path,
                    file_size=file_size,
                    format=ExportFormat.ZIP,
                    export_type=ExportType.SESSION_DATA,
                    records_exported=len(package_files),
                    metadata={"package_files": [f.name for f in package_files if f]},
                )

        except Exception as e:
            logger.error(f"Erro ao criar pacote de exportação: {e}")
            return ExportResult(
                success=False,
                format=ExportFormat.ZIP,
                export_type=ExportType.SESSION_DATA,
                error_message=str(e),
            )

    def _emit_export_event(self, result: ExportResult, session_id: str) -> None:
        """Disparar evento de exportação."""
        try:
            event = DataEvent(
                source="export_import_manager",
                data_type="export",
                data_size=result.file_size,
                metadata={
                    "action": "export_completed",
                    "session_id": session_id,
                    "format": result.format.value if result.format else "unknown",
                    "export_type": result.export_type.value if result.export_type else "unknown",
                    "success": result.success,
                    "records_exported": result.records_exported,
                    "duration": result.duration,
                    "file_path": str(result.file_path) if result.file_path else None,
                },
            )

            event_bus.publish_sync(event)
        except Exception as e:
            logger.debug(f"Erro ao disparar evento de exportação: {e}")

    def _emit_import_event(self, result: ImportResult, file_path: str) -> None:
        """Disparar evento de importação."""
        try:
            event = DataEvent(
                source="export_import_manager",
                data_type="import",
                metadata={
                    "action": "import_completed",
                    "file_path": file_path,
                    "format": result.format.value if result.format else "unknown",
                    "success": result.success,
                    "records_imported": result.records_imported,
                    "duration": result.duration,
                },
            )

            event_bus.publish_sync(event)
        except Exception as e:
            logger.debug(f"Erro ao disparar evento de importação: {e}")

    def get_available_formats(self) -> Dict[str, List[str]]:
        """Obter formatos disponíveis."""
        return {
            "export": [fmt.value for fmt in self.exporters.keys()],
            "import": [fmt.value for fmt in self.importers.keys()],
        }

    def get_export_history(self, limit: int = 10) -> List[Path]:
        """Obter histórico de arquivos exportados."""
        try:
            export_files = list(self.default_export_dir.glob("*"))
            export_files.sort(key=lambda x: x.stat().st_mtime, reverse=True)
            return export_files[:limit]
        except Exception as e:
            logger.error(f"Erro ao obter histórico de exportação: {e}")
            return []

    def cleanup_old_exports(self, days: int = 7) -> int:
        """Limpar exportações antigas."""
        try:
            cutoff_time = datetime.now().timestamp() - (days * 24 * 3600)
            removed_count = 0

            for file_path in self.default_export_dir.glob("*"):
                if file_path.stat().st_mtime < cutoff_time:
                    try:
                        file_path.unlink()
                        removed_count += 1
                    except Exception as e:
                        logger.warning(f"Erro ao remover arquivo {file_path}: {e}")

            logger.info(f"Removidos {removed_count} arquivos de exportação antigos")
            return removed_count

        except Exception as e:
            logger.error(f"Erro na limpeza de exportações: {e}")
            return 0


# Instância global do gerenciador
export_import_manager = ExportImportManager()
